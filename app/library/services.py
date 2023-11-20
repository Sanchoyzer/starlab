import io
from collections.abc import Sequence
from typing import Self
from uuid import UUID

import openpyxl
from sqlalchemy import or_, select, update

from app.library.database import async_session
from app.library.exceptions import BookImportError
from app.library.models import Book
from app.library.schemas import BookCreateRequest, BookListFilters


class BookService:
    @classmethod
    async def create(cls: type[Self], book_in: BookCreateRequest) -> Book:
        async with async_session() as session, session.begin():
            book = Book(**book_in.model_dump())
            session.add(book)
            return book

    @classmethod
    async def get_one(cls: type[Self], uid: UUID) -> Book | None:
        async with async_session() as session:
            query = select(Book).filter_by(uid=uid)
            result = await session.execute(query)
            return result.scalars().first()

    @classmethod
    async def get_list(cls: type[Self], filters: BookListFilters) -> Sequence[Book]:
        async with async_session() as session:
            query = select(Book)
            if name := filters.name:
                query = query.filter(Book.name.ilike(f'%{name}%'))
            if author := filters.author:
                query = query.filter(Book.author.ilike(f'%{author}%'))
            if published_at := filters.published_at:
                query = query.filter_by(published_at=published_at)
            if genre := filters.genre:
                query = query.filter(Book.genre.in_(genre))
            query = query.limit(limit=filters.limit).offset(offset=filters.offset)
            result = await session.execute(query)
            return result.scalars().all()

    @classmethod
    async def import_books(cls: type[Self], data: io.BytesIO) -> None:
        try:
            workbook = openpyxl.load_workbook(data)
            sheet_names = workbook.sheetnames

            names = set()
            if 'name' in sheet_names:
                for row in workbook['name'].iter_rows(min_row=2, min_col=2, max_col=2):
                    value = str(int(row[0].value)) if row[0].data_type == 'n' else row[0].value
                    names.add(value)

            authors = set()
            if 'author' in sheet_names:
                for row in workbook['author'].iter_rows(min_row=2, min_col=2, max_col=2):
                    authors.add(row[0].value)

            if not names and not authors:
                raise ValueError('empty excel file')  # noqa: TRY301
        except Exception as exc:  # noqa: BLE001
            raise BookImportError(repr(exc)) from exc

        async with async_session() as session:
            query = (
                update(Book)
                .where(or_(Book.name.in_(names), Book.author.in_(authors)))
                .values(available_for_download=False)
            )
            await session.execute(query)
            await session.commit()
